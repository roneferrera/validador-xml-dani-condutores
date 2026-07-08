import streamlit as st
import pandas as pd
import io
import re
import zipfile
from lxml import etree

st.set_page_config(
    page_title="Enriquecedor de NF-e — DNI",
    page_icon="🟠",
    layout="wide",
    initial_sidebar_state="expanded",
)

CATALOGO_PADRAO: dict[int, str] = {
    1201: "Dev. de Venda de Prod. do Estabelecimento — Estadual",
    1202: "Dev. de Venda de Merc. Adq./Rec. de Terceiros — Estadual",
    1410: "Devolução de Venda Própria ST — Estadual",
    1411: "Devolução de Venda Terceiros ST — Estadual",
    1949: "Outra Entrada de Mercadoria — Estadual",
    2201: "Dev. de Venda de Prod. do Estabelecimento — Interestadual",
    2202: "Dev. de Venda de Merc. Adq./Rec. de Terceiros — Interestadual",
    2410: "Devolução de Venda Própria ST — Interestadual",
    2411: "Devolução de Venda Terceiros ST — Interestadual",
    2603: "Devolução de Venda de Energia Elétrica — Interestadual",
    2949: "Outra Entrada de Mercadoria — Interestadual",
}
ATIVOS_PADRAO: set[int] = set(CATALOGO_PADRAO.keys())

st.markdown("""
<style>
html, body, [class*="css"] { font-family: 'Segoe UI', 'Arial', sans-serif; color: #444444; }
h1, h2, h3 { color: #FF8000; font-weight: 700; }
.main-header { background: #444444; padding: 18px 28px 14px 28px; border-radius: 8px; border-top: 6px solid #FF8000; margin-bottom: 20px; }
.main-header h2 { color: #FF8000 !important; margin: 0; font-size: 1.4rem; }
.main-header p  { color: #DDDDDD !important; margin: 4px 0 0 0; font-size: 0.82rem; }
section[data-testid="stSidebar"] { background-color: #2E2E2E; }
section[data-testid="stSidebar"] * { color: #FFFFFF !important; }
section[data-testid="stSidebar"] .stButton > button { background-color: #FF8000 !important; color: #FFFFFF !important; border: none !important; border-radius: 4px !important; font-weight: bold !important; }
section[data-testid="stSidebar"] .stButton > button:hover { background-color: #D64001 !important; }
.stButton > button { background-color: #FF8000 !important; color: #FFFFFF !important; border: none !important; border-radius: 4px !important; font-weight: bold !important; }
.stButton > button:hover { background-color: #D64001 !important; }
.stDownloadButton > button { background-color: #FF8000 !important; color: #FFFFFF !important; border: none !important; border-radius: 4px !important; font-weight: bold !important; }
.stDownloadButton > button:hover { background-color: #D64001 !important; }
[data-testid="metric-container"] { background-color: #F5F5F5; border-left: 4px solid #FF8000; border-radius: 4px; padding: 8px 12px; }
.footer { text-align: center; color: #999; font-size: 0.72rem; margin-top: 24px; padding-top: 12px; border-top: 1px solid #E0E0E0; }
</style>
""", unsafe_allow_html=True)

# ── Session State ──────────────────────────────────────────────────────────────
def _init_state():
    defaults = {
        "cfop_catalogo": dict(CATALOGO_PADRAO),
        "cfop_ativos": set(ATIVOS_PADRAO),
        "resultado_excel_bytes": None,
        "resultado_xmls_modificados": {},
        "resultado_xmls_originais": {},
        "resultado_log": [],
        "resultado_metricas": None,
        "resultado_diferencas": [],
        "processamento_concluido": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

def _limpar_resultados():
    st.session_state.resultado_excel_bytes = None
    st.session_state.resultado_xmls_modificados = {}
    st.session_state.resultado_xmls_originais = {}
    st.session_state.resultado_log = []
    st.session_state.resultado_metricas = None
    st.session_state.resultado_diferencas = []
    st.session_state.processamento_concluido = False

# ── Sidebar ────────────────────────────────────────────────────────────────────
def _extrair_cfops_texto(texto: str) -> list[int]:
    return [int(t) for t in re.findall(r"\d+", texto) if len(t) == 4]

def _build_rows() -> list[dict]:
    catalogo = st.session_state.cfop_catalogo
    ativos   = st.session_state.cfop_ativos
    return [{"Ativo": cod in ativos, "Código": cod, "Descrição": desc}
            for cod, desc in sorted(catalogo.items())]

def render_sidebar():
    with st.sidebar:
        st.markdown("## ⚙️ Configuração de CFOPs")
        st.markdown("---")
        st.markdown("### 📥 CFOPs de Devolução (Planilha)")
        st.markdown(
            "<small>CFOPs de <b>entrada</b> da planilha (DNI) usados para filtrar devoluções. "
            "O CFOP original do XML <b>não é alterado</b>.</small>",
            unsafe_allow_html=True,
        )
        texto = st.text_area("CFOPs em lote", placeholder="Ex:\n1201, 1410\n2411",
                             height=110, key="sidebar_cfop_texto", label_visibility="collapsed")
        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("➕ Ativar", use_container_width=True, key="btn_ativar_lote"):
                codigos = _extrair_cfops_texto(texto)
                if codigos:
                    for c in codigos:
                        if c not in st.session_state.cfop_catalogo:
                            st.session_state.cfop_catalogo[c] = "CFOP Personalizado"
                        st.session_state.cfop_ativos.add(c)
                    st.toast(f"✅ {len(codigos)} CFOP(s) ativado(s).")
                    st.rerun()
                else:
                    st.warning("Nenhum CFOP válido encontrado.")
        with col_b:
            if st.button("🗑️ Desativar", use_container_width=True, key="btn_desativar_lote"):
                codigos = _extrair_cfops_texto(texto)
                if codigos:
                    for c in codigos: st.session_state.cfop_ativos.discard(c)
                    st.toast("🗑️ CFOPs desativados.")
                    st.rerun()
                else:
                    st.warning("Nenhum CFOP válido encontrado.")

        st.markdown("---")
        if st.button("🔄 Redefinir para o Padrão", use_container_width=True, key="btn_reset_cfop"):
            st.session_state.cfop_catalogo = dict(CATALOGO_PADRAO)
            st.session_state.cfop_ativos   = set(ATIVOS_PADRAO)
            st.toast("✅ Catálogo restaurado.")
            st.rerun()
        if st.button("🔲 Desmarcar Todos", use_container_width=True, key="btn_desmarcar"):
            st.session_state.cfop_ativos = set()
            st.toast("🔲 Todos desmarcados.")
            st.rerun()

        st.markdown("---")
        st.markdown("### 🗂️ CFOPs Cadastrados")
        edited = st.data_editor(
            _build_rows(),
            column_config={
                "Ativo":     st.column_config.CheckboxColumn("Ativo", default=False),
                "Código":    st.column_config.NumberColumn("Código", format="%d", disabled=True),
                "Descrição": st.column_config.TextColumn("Descrição", disabled=True),
            },
            hide_index=True, use_container_width=True, key="cfop_data_editor",
        )
        novos_ativos: set[int] = {r["Código"] for r in edited if r["Ativo"]}
        if novos_ativos != st.session_state.cfop_ativos:
            st.session_state.cfop_ativos = novos_ativos
            st.rerun()

        st.markdown("---")
        st.markdown("#### 🐍 CFOPs ativos (backend)")
        st.code(str(sorted(st.session_state.cfop_ativos)), language="python")
        total  = len(st.session_state.cfop_catalogo)
        ativos = len(st.session_state.cfop_ativos)
        st.caption(f"Catálogo: {total} | Ativos: {ativos} | Inativos: {total - ativos}")
        st.markdown("---")
        st.markdown("**Thomson Reuters · Domínio Sistemas**")
        st.markdown("**Enriquecedor NF-e v8.0**")

# ── Helpers gerais ─────────────────────────────────────────────────────────────
def limpar_chave(valor: str) -> str:
    return re.sub(r"[^0-9]", "", str(valor).strip()) if valor else ""

def limpar_valor(valor) -> float:
    if valor is None: return 0.0
    try:
        if pd.isna(valor): return 0.0
    except Exception: pass
    s = re.sub(r"[^\d.\-]", "", str(valor).strip().replace(",", "."))
    try: return float(s)
    except ValueError: return 0.0

def fmt(v: float) -> str:
    return f"{v:.2f}"

def safe_int_cst(val) -> str:
    if val is None: return "00"
    try:
        if pd.isna(val): return "00"
    except Exception: pass
    s = str(val).strip()
    if s.lower() in ("nan", "none", ""): return "00"
    try: return str(int(float(s))).zfill(2)
    except Exception: return "00"

# ── Detecção automática de encoding ───────────────────────────────────────────
def detectar_encoding(conteudo_bytes: bytes) -> str:
    """
    Detecta o encoding do arquivo em cascata:
      1. BOM explícito (UTF-32, UTF-16, UTF-8 BOM)
      2. chardet com confiança >= 70% (opcional — não quebra se não instalado)
      3. Tentativas manuais em ordem de prevalência (BR)
      4. Fallback final: latin-1 (nunca falha)
    """
    # ── 1. Detecção por BOM ────────────────────────────────────────────────────
    BOMS = [
        (b"\xFF\xFE\x00\x00", "utf-32-le"),
        (b"\x00\x00\xFE\xFF", "utf-32-be"),
        (b"\xFF\xFE",         "utf-16-le"),
        (b"\xFE\xFF",         "utf-16-be"),
        (b"\xEF\xBB\xBF",    "utf-8-sig"),   # UTF-8 com BOM — Excel "Salvar como CSV"
    ]
    for bom, enc in BOMS:
        if conteudo_bytes.startswith(bom):
            return enc

    # ── 2. chardet (análise estatística) ──────────────────────────────────────
    try:
        import chardet
        amostra    = conteudo_bytes[:32_768]
        resultado  = chardet.detect(amostra)
        enc_bruto  = (resultado.get("encoding") or "").lower().replace("-", "").replace("_", "")
        confianca  = resultado.get("confidence") or 0

        ALIAS = {
            "ascii":       "utf-8",
            "utf8":        "utf-8",
            "utf8sig":     "utf-8-sig",
            "iso88591":    "latin-1",
            "iso88592":    "latin-1",
            "windows1252": "cp1252",
            "windows1250": "cp1250",
        }
        enc_detectado = ALIAS.get(enc_bruto, enc_bruto)

        if enc_detectado and confianca >= 0.70:
            try:
                conteudo_bytes.decode(enc_detectado)
                return enc_detectado
            except (UnicodeDecodeError, LookupError):
                pass
    except ImportError:
        pass  # chardet não instalado — segue para tentativas manuais

    # ── 3. Tentativas manuais em ordem de prevalência (Brasil) ────────────────
    CANDIDATOS = ["utf-8", "cp1252", "latin-1", "cp1250", "utf-16"]
    for enc in CANDIDATOS:
        try:
            conteudo_bytes.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue

    # ── 4. Fallback absoluto — latin-1 nunca lança UnicodeDecodeError ─────────
    return "latin-1"

# ── Leitura XLSX / CSV ─────────────────────────────────────────────────────────
def ler_planilha(conteudo_bytes: bytes, nome_arquivo: str) -> dict:
    indexado = {}
    try:
        nome_lower = nome_arquivo.lower()
        if nome_lower.endswith(".csv"):
            # ── Detecta encoding automaticamente ──────────────────────────────
            encoding = detectar_encoding(conteudo_bytes)

            # ── Detecta separador (vírgula ou ponto-e-vírgula) ─────────────────
            try:
                amostra = conteudo_bytes[:4096].decode(encoding, errors="replace")
            except Exception:
                amostra = conteudo_bytes[:4096].decode("latin-1", errors="replace")
            sep = ";" if amostra.count(";") >= amostra.count(",") else ","

            # ── Tenta ler com o encoding detectado; fallback em cascata ────────
            FALLBACKS = list(dict.fromkeys([encoding, "cp1252", "utf-8", "latin-1"]))
            df = None
            ultimo_erro = None
            for enc in FALLBACKS:
                try:
                    df = pd.read_csv(
                        io.BytesIO(conteudo_bytes),
                        dtype=str,
                        sep=sep,
                        encoding=enc,
                        on_bad_lines="skip",
                    )
                    encoding = enc   # registra o que funcionou
                    break
                except (UnicodeDecodeError, Exception) as e:
                    ultimo_erro = e
                    continue

            if df is None:
                st.error(f"Erro ao ler CSV após todas as tentativas de encoding: {ultimo_erro}")
                return indexado

            st.caption(f"📄 Encoding detectado: `{encoding}` | Separador: `{sep}` | Linhas: `{len(df)}`")

        else:
            # XLSX/XLS — openpyxl gerencia encoding internamente
            df = pd.read_excel(io.BytesIO(conteudo_bytes), dtype=str, engine="openpyxl")

    except Exception as e:
        st.error(f"Erro ao ler planilha: {e}")
        return indexado

    df.columns = [str(c).strip() for c in df.columns]

    for _, row in df.iterrows():
        reg = {k: (str(v).strip() if pd.notna(v) else "") for k, v in row.items()}
        chave_nfe = limpar_chave(reg.get("Chave Nfe/Cte", ""))
        seq_raw   = reg.get("Sequencia", "0")
        try:   seq = int(float(seq_raw)) if seq_raw else 0
        except Exception: seq = 0
        if len(chave_nfe) < 44 or seq == 0: continue

        cfop_xlsx     = reg.get("Cfop", "").strip()
        vlr_documento = limpar_valor(reg.get("Vlr Documento", "0"))
        vlr_icms      = limpar_valor(reg.get("Vlr Icms", "0"))
        vlr_icms_st   = limpar_valor(reg.get("Vlr Icms St", "0"))
        vlr_ipi       = limpar_valor(reg.get("Vlr Ipi", "0"))
        bc_pis_cofins = max(0.0, vlr_documento - vlr_icms - vlr_icms_st - vlr_ipi)

        cst_icms = safe_int_cst(reg.get("CST ICMS", ""))
        if cst_icms == "00" and (
            limpar_valor(reg.get("Base Icms St", "0")) > 0 or
            limpar_valor(reg.get("Vlr Icms St",  "0")) > 0
        ):
            cst_icms = "10"

        base_ipi = limpar_valor(reg.get("Base Ipi", "0"))
        perc_ipi = limpar_valor(reg.get("Perc Ipi", "0"))
        vlr_ipi  = limpar_valor(reg.get("Vlr Ipi",  "0"))
        tem_ipi  = (base_ipi > 0 or vlr_ipi > 0)

        cst_ipi_raw = safe_int_cst(reg.get("CST IPI", ""))
        if tem_ipi and cst_ipi_raw in ("00", ""):
            cst_ipi_final = "50"
        else:
            cst_ipi_final = cst_ipi_raw

        indexado[(chave_nfe, seq)] = {
            "cfop_xlsx":      cfop_xlsx,
            "cod_item":       reg.get("Cod Item", "").strip(),
            "ncm":            reg.get("NCM", "").strip(),
            "nro_documento":  reg.get("Nro Documento", "").strip(),
            "razao_social":   reg.get("Razao Social", "").strip(),
            "cnpj":           reg.get("CNPJ-CPF", "").strip(),
            "data_entrada":   reg.get("Data Entrada", "").strip(),
            "data_emissao":   reg.get("Data Emissao", "").strip(),
            "uf":             reg.get("UF", "").strip(),
            "cst_icms":       cst_icms,
            "base_icms":      limpar_valor(reg.get("Base Icms", "0")),
            "perc_icms":      limpar_valor(reg.get("Perc ICms", "0")),
            "vlr_icms":       vlr_icms,
            "base_icms_st":   limpar_valor(reg.get("Base Icms St", "0")),
            "vlr_icms_st":    vlr_icms_st,
            "cst_ipi":        cst_ipi_final,
            "base_ipi":       base_ipi,
            "perc_ipi":       perc_ipi,
            "vlr_ipi":        vlr_ipi,
            "tem_ipi":        tem_ipi,
            "cst_pis":        safe_int_cst(reg.get("CST PIS", "")),
            "perc_pis":       limpar_valor(reg.get("Perc Pis", "0")),
            "vlr_pis":        limpar_valor(reg.get("Vlr Pis", "0")),
            "cst_cofins":     safe_int_cst(reg.get("CST COFINS", "")),
            "perc_cofins":    limpar_valor(reg.get("Perc Cofins", "0")),
            "vlr_cofins":     limpar_valor(reg.get("Vlr Cofins", "0")),
            "bc_pis_cofins":  bc_pis_cofins,
            "vlr_documento":  vlr_documento,
        }
    return indexado

# ── Helpers XML ────────────────────────────────────────────────────────────────
NS = "http://www.portalfiscal.inf.br/nfe"
def nstag(n):  return f"{{{NS}}}{n}"
def local(e):  return e.tag.split("}")[-1] if "}" in e.tag else e.tag
def find(elem, *nomes):
    atual = elem
    for nome in nomes:
        enc = next((f for f in atual if local(f) == nome), None)
        if enc is None: return None
        atual = enc
    return atual

def _insert_after(parent, ref_tag: str, new_elem):
    for i, child in enumerate(parent):
        if local(child) == ref_tag:
            parent.insert(i + 1, new_elem)
            return
    parent.append(new_elem)

# ── ICMS ───────────────────────────────────────────────────────────────────────
CST_COM_ST = {"10", "30", "70", "90"}

def aplicar_icms(icms_filho, dados):
    modificado = False
    cst = dados["cst_icms"]
    el = find(icms_filho, "CST")
    if el is not None: el.text = cst; modificado = True
    for tn, tv in [("vBC", fmt(dados["base_icms"])), ("pICMS", fmt(dados["perc_icms"])), ("vICMS", fmt(dados["vlr_icms"]))]:
        el = find(icms_filho, tn)
        if el is not None: el.text = tv; modificado = True
    tem_st = cst in CST_COM_ST and (dados["base_icms_st"] > 0 or dados["vlr_icms_st"] > 0)
    if tem_st:
        el_mod = find(icms_filho, "modBCST")
        if el_mod is None:
            idx = next((i for i, f in enumerate(icms_filho) if local(f) == "vICMS"), None)
            el_mod = etree.Element(nstag("modBCST")); el_mod.text = "4"
            icms_filho.insert(idx + 1, el_mod) if idx is not None else icms_filho.append(el_mod)
        if (el_mod.text or "4").strip() == "4":
            el_pmva = find(icms_filho, "pMVAST")
            if el_pmva is None:
                idx = next((i for i, f in enumerate(icms_filho) if local(f) == "modBCST"), None)
                el_pmva = etree.Element(nstag("pMVAST")); el_pmva.text = "0.00"
                icms_filho.insert(idx + 1, el_pmva) if idx is not None else icms_filho.append(el_pmva)
        for tn, tv in [("vBCST", fmt(dados["base_icms_st"])), ("pICMSST", fmt(dados["perc_icms"])), ("vICMSST", fmt(max(0.0, dados["vlr_icms_st"])))]:
            el = find(icms_filho, tn)
            if el is not None: el.text = tv; modificado = True
            else:
                novo = etree.SubElement(icms_filho, nstag(tn)); novo.text = tv; modificado = True
    return modificado

# ── IPI + impostoDevol ─────────────────────────────────────────────────────────
def aplicar_ipi(imposto_elem, det_elem, dados):
    modificado = False
    tem_ipi    = dados.get("tem_ipi", False)
    cst_ipi    = dados["cst_ipi"]

    if tem_ipi:
        ipi_elem = find(imposto_elem, "IPI")

        if ipi_elem is None:
            ipi_elem = etree.Element(nstag("IPI"))
            el_cenq  = etree.SubElement(ipi_elem, nstag("cEnq")); el_cenq.text = "999"
            ipi_trib = etree.SubElement(ipi_elem, nstag("IPITrib"))
            etree.SubElement(ipi_trib, nstag("CST")).text  = cst_ipi
            etree.SubElement(ipi_trib, nstag("vBC")).text  = fmt(dados["base_ipi"])
            etree.SubElement(ipi_trib, nstag("pIPI")).text = fmt(dados["perc_ipi"])
            etree.SubElement(ipi_trib, nstag("vIPI")).text = fmt(dados["vlr_ipi"])
            pis_elem = find(imposto_elem, "PIS")
            if pis_elem is not None:
                imposto_elem.insert(list(imposto_elem).index(pis_elem), ipi_elem)
            else:
                imposto_elem.append(ipi_elem)
            modificado = True

        else:
            ipint    = find(ipi_elem, "IPINT")
            ipi_trib = find(ipi_elem, "IPITrib")

            if ipint is not None:
                ipi_elem.remove(ipint)
                ipi_trib = etree.SubElement(ipi_elem, nstag("IPITrib"))
                etree.SubElement(ipi_trib, nstag("CST")).text  = cst_ipi
                etree.SubElement(ipi_trib, nstag("vBC")).text  = fmt(dados["base_ipi"])
                etree.SubElement(ipi_trib, nstag("pIPI")).text = fmt(dados["perc_ipi"])
                etree.SubElement(ipi_trib, nstag("vIPI")).text = fmt(dados["vlr_ipi"])
                modificado = True

            elif ipi_trib is not None:
                el_cst = find(ipi_trib, "CST")
                if el_cst is not None:
                    el_cst.text = cst_ipi; modificado = True

                for tag_rem in ["qUnid", "vUnid"]:
                    el_rem = find(ipi_trib, tag_rem)
                    if el_rem is not None:
                        ipi_trib.remove(el_rem); modificado = True

                el_vbc = find(ipi_trib, "vBC")
                if el_vbc is None:
                    el_vbc = etree.Element(nstag("vBC"))
                    _insert_after(ipi_trib, "CST", el_vbc)
                el_vbc.text = fmt(dados["base_ipi"]); modificado = True

                el_pipi = find(ipi_trib, "pIPI")
                if el_pipi is None:
                    el_pipi = etree.Element(nstag("pIPI"))
                    _insert_after(ipi_trib, "vBC", el_pipi)
                el_pipi.text = fmt(dados["perc_ipi"]); modificado = True

                el_vipi = find(ipi_trib, "vIPI")
                if el_vipi is None:
                    el_vipi = etree.Element(nstag("vIPI"))
                    _insert_after(ipi_trib, "pIPI", el_vipi)
                el_vipi.text = fmt(dados["vlr_ipi"]); modificado = True

    # Sempre zera impostoDevol se existir
    imp_devol = find(det_elem, "impostoDevol")
    if imp_devol is not None:
        el_pdevol = find(imp_devol, "pDevol")
        if el_pdevol is not None: el_pdevol.text = "0.00"; modificado = True
        ipi_devol = find(imp_devol, "IPI")
        if ipi_devol is not None:
            el_vipidevol = find(ipi_devol, "vIPIDevol")
            if el_vipidevol is not None: el_vipidevol.text = "0.00"; modificado = True

    return modificado

# ── Processamento XML ──────────────────────────────────────────────────────────
def processar_xml(conteudo_xml, nome_arquivo, dados_indexados, cfops_ativas_xlsx):
    try: tree = etree.fromstring(conteudo_xml)
    except Exception as e: return None, f"XML inválido: {e}", "erro", []

    inf_nfe = find(tree, "NFe", "infNFe") or find(tree, "infNFe")
    if inf_nfe is None: return None, "infNFe não encontrado", "erro", []

    chave_xml = re.sub(r"[^0-9]", "", inf_nfe.get("Id", ""))
    if len(chave_xml) < 44: return None, f"Chave inválida: {chave_xml}", "erro", []

    det_elements = [f for f in inf_nfe if local(f) == "det"]
    itens_validos = []
    for det in det_elements:
        try:   n_item = int(det.get("nItem", "0"))
        except Exception: continue
        dados = dados_indexados.get((chave_xml, n_item))
        if dados is None: continue
        if dados.get("cfop_xlsx", "") in cfops_ativas_xlsx:
            itens_validos.append((det, n_item, dados))

    if not itens_validos:
        return None, "nenhum item com CFOP válido — não alterado", "info", []

    modificado = False
    diferencas = []

    for det, n_item, dados in itens_validos:
        prod    = find(det, "prod")
        imposto = find(det, "imposto")
        if prod is None or imposto is None: continue

        def _get(elem, *path):
            el = find(elem, *path)
            return (el.text or "0").strip() if el is not None else "0"

        def _fv(v):
            try: return round(float(str(v).replace(",", ".")), 2)
            except Exception: return 0.0

        cfop_xml_original = _get(prod, "CFOP")

        # ── Captura estado ANTES ──────────────────────────────────────────────
        antes = {}
        icms_pai = find(imposto, "ICMS")
        if icms_pai:
            for icms_f in icms_pai:
                if local(icms_f).startswith("ICMS"):
                    antes["CST ICMS"]    = _get(icms_f, "CST")
                    antes["BC ICMS"]     = _get(icms_f, "vBC")
                    antes["% ICMS"]      = _get(icms_f, "pICMS")
                    antes["Vlr ICMS"]    = _get(icms_f, "vICMS")
                    antes["BC ICMS ST"]  = _get(icms_f, "vBCST")
                    antes["Vlr ICMS ST"] = _get(icms_f, "vICMSST")
                    break

        ipi_elem_antes = find(imposto, "IPI")
        if ipi_elem_antes:
            ipi_trib_antes = find(ipi_elem_antes, "IPITrib")
            ipi_nt_antes   = find(ipi_elem_antes, "IPINT")
            if ipi_trib_antes:
                antes["CST IPI"] = _get(ipi_trib_antes, "CST")
                vbc_antes = _get(ipi_trib_antes, "vBC")
                if vbc_antes == "0": vbc_antes = _get(ipi_trib_antes, "qUnid")
                antes["BC IPI"]  = vbc_antes
                antes["% IPI"]   = _get(ipi_trib_antes, "pIPI")
                antes["Vlr IPI"] = _get(ipi_trib_antes, "vIPI")
            elif ipi_nt_antes:
                antes["CST IPI"] = _get(ipi_nt_antes, "CST")
                antes["BC IPI"]  = "0"
                antes["% IPI"]   = "0"
                antes["Vlr IPI"] = "0"

        pis_p = find(imposto, "PIS")
        if pis_p:
            for pf in pis_p:
                antes["CST PIS"] = _get(pf, "CST"); antes["BC PIS"] = _get(pf, "vBC")
                antes["% PIS"]   = _get(pf, "pPIS"); antes["Vlr PIS"] = _get(pf, "vPIS"); break
        cof_p = find(imposto, "COFINS")
        if cof_p:
            for cf in cof_p:
                antes["CST COFINS"] = _get(cf, "CST"); antes["BC COFINS"] = _get(cf, "vBC")
                antes["% COFINS"]   = _get(cf, "pCOFINS"); antes["Vlr COFINS"] = _get(cf, "vCOFINS"); break
        imp_devol_el       = find(det, "impostoDevol")
        antes["pDevol"]    = _get(imp_devol_el, "pDevol")           if imp_devol_el is not None else "0"
        antes["vIPIDevol"] = _get(imp_devol_el, "IPI", "vIPIDevol") if imp_devol_el is not None else "0"

        # ── Aplica alterações ─────────────────────────────────────────────────
        el_ncm = find(prod, "NCM")
        if el_ncm is not None and dados["ncm"]: el_ncm.text = dados["ncm"]; modificado = True

        if icms_pai is not None:
            for icms_f in icms_pai:
                if local(icms_f).startswith("ICMS"):
                    if aplicar_icms(icms_f, dados): modificado = True
                    break

        if aplicar_ipi(imposto, det, dados): modificado = True

        pis_pai = find(imposto, "PIS")
        if pis_pai is not None:
            for pf in pis_pai:
                for tn, tv in [("CST", dados["cst_pis"]), ("vBC", fmt(dados["bc_pis_cofins"])),
                               ("pPIS", fmt(dados["perc_pis"])), ("vPIS", fmt(dados["vlr_pis"]))]:
                    el = find(pf, tn)
                    if el is not None: el.text = tv; modificado = True
                break

        cof_pai = find(imposto, "COFINS")
        if cof_pai is not None:
            for cf in cof_pai:
                for tn, tv in [("CST", dados["cst_cofins"]), ("vBC", fmt(dados["bc_pis_cofins"])),
                               ("pCOFINS", fmt(dados["perc_cofins"])), ("vCOFINS", fmt(dados["vlr_cofins"]))]:
                    el = find(cf, tn)
                    if el is not None: el.text = tv; modificado = True
                break

        # ── Estado DEPOIS ─────────────────────────────────────────────────────
        depois = {
            "CST ICMS": dados["cst_icms"], "BC ICMS": dados["base_icms"],
            "% ICMS": dados["perc_icms"],  "Vlr ICMS": dados["vlr_icms"],
            "BC ICMS ST": dados["base_icms_st"], "Vlr ICMS ST": max(0.0, dados["vlr_icms_st"]),
            "CST IPI": dados["cst_ipi"],   "BC IPI": dados["base_ipi"],
            "% IPI": dados["perc_ipi"],    "Vlr IPI": dados["vlr_ipi"],
            "CST PIS": dados["cst_pis"],   "BC PIS": dados["bc_pis_cofins"],
            "% PIS": dados["perc_pis"],    "Vlr PIS": dados["vlr_pis"],
            "CST COFINS": dados["cst_cofins"], "BC COFINS": dados["bc_pis_cofins"],
            "% COFINS": dados["perc_cofins"],  "Vlr COFINS": dados["vlr_cofins"],
            "pDevol": 0.0, "vIPIDevol": 0.0,
        }

        campos_num = ["BC ICMS","Vlr ICMS","BC ICMS ST","Vlr ICMS ST",
                      "BC IPI","Vlr IPI","BC PIS","Vlr PIS","BC COFINS","Vlr COFINS","vIPIDevol"]
        tem_diff = any(
            round(_fv(antes.get(k, "0")), 2) != round(float(depois.get(k, 0)), 2)
            for k in campos_num
        )

        xprod_el = find(prod, "xProd")
        xprod    = xprod_el.text if xprod_el is not None else ""

        row_conf = {
            "Arquivo":                   nome_arquivo,
            "Chave NF-e":                chave_xml,
            "Nro Documento":             dados["nro_documento"],
            "Data Emissão":              dados["data_emissao"],
            "Data Entrada":              dados["data_entrada"],
            "Razão Social":              dados["razao_social"],
            "CNPJ/CPF":                  dados["cnpj"],
            "UF":                        dados["uf"],
            "nItem":                     n_item,
            "Cód Item":                  dados["cod_item"],
            "Desc Item":                 xprod,
            "NCM":                       dados["ncm"],
            "CFOP XML (origem)":         cfop_xml_original,
            "CFOP XLSX (DNI/entrada)":   dados["cfop_xlsx"],
            "Vlr Documento":             round(dados["vlr_documento"], 2),
            "CST ICMS Antes":            antes.get("CST ICMS", ""),
            "CST ICMS Depois":           depois["CST ICMS"],
            "BC ICMS Antes":             round(_fv(antes.get("BC ICMS", "0")), 2),
            "BC ICMS Depois":            round(float(depois["BC ICMS"]), 2),
            "% ICMS":                    round(float(depois["% ICMS"]), 4),
            "Vlr ICMS Antes":            round(_fv(antes.get("Vlr ICMS", "0")), 2),
            "Vlr ICMS Depois":           round(float(depois["Vlr ICMS"]), 2),
            "Diff Vlr ICMS":             round(float(depois["Vlr ICMS"]) - _fv(antes.get("Vlr ICMS","0")), 2),
            "BC ICMS ST Antes":          round(_fv(antes.get("BC ICMS ST", "0")), 2),
            "BC ICMS ST Depois":         round(float(depois["BC ICMS ST"]), 2),
            "Vlr ICMS ST Antes":         round(_fv(antes.get("Vlr ICMS ST", "0")), 2),
            "Vlr ICMS ST Depois":        round(float(depois["Vlr ICMS ST"]), 2),
            "Diff Vlr ICMS ST":          round(float(depois["Vlr ICMS ST"]) - _fv(antes.get("Vlr ICMS ST","0")), 2),
            "CST IPI Antes":             antes.get("CST IPI", ""),
            "CST IPI Depois":            depois["CST IPI"],
            "BC IPI Antes":              round(_fv(antes.get("BC IPI", "0")), 2),
            "BC IPI Depois":             round(float(depois["BC IPI"]), 2),
            "% IPI":                     round(float(depois["% IPI"]), 4),
            "Vlr IPI Antes":             round(_fv(antes.get("Vlr IPI", "0")), 2),
            "Vlr IPI Depois":            round(float(depois["Vlr IPI"]), 2),
            "Diff Vlr IPI":              round(float(depois["Vlr IPI"]) - _fv(antes.get("Vlr IPI","0")), 2),
            "pDevol Antes":              round(_fv(antes.get("pDevol", "0")), 2),
            "pDevol Depois":             0.00,
            "vIPIDevol Antes":           round(_fv(antes.get("vIPIDevol", "0")), 2),
            "vIPIDevol Depois":          0.00,
            "Diff vIPIDevol":            round(0.0 - _fv(antes.get("vIPIDevol","0")), 2),
            "CST PIS Antes":             antes.get("CST PIS", ""),
            "CST PIS Depois":            depois["CST PIS"],
            "BC PIS Antes":              round(_fv(antes.get("BC PIS", "0")), 2),
            "BC PIS Depois":             round(float(depois["BC PIS"]), 2),
            "% PIS":                     round(float(depois["% PIS"]), 4),
            "Vlr PIS Antes":             round(_fv(antes.get("Vlr PIS", "0")), 2),
            "Vlr PIS Depois":            round(float(depois["Vlr PIS"]), 2),
            "Diff Vlr PIS":              round(float(depois["Vlr PIS"]) - _fv(antes.get("Vlr PIS","0")), 2),
            "CST COFINS Antes":          antes.get("CST COFINS", ""),
            "CST COFINS Depois":         depois["CST COFINS"],
            "BC COFINS Antes":           round(_fv(antes.get("BC COFINS", "0")), 2),
            "BC COFINS Depois":          round(float(depois["BC COFINS"]), 2),
            "% COFINS":                  round(float(depois["% COFINS"]), 4),
            "Vlr COFINS Antes":          round(_fv(antes.get("Vlr COFINS", "0")), 2),
            "Vlr COFINS Depois":         round(float(depois["Vlr COFINS"]), 2),
            "Diff Vlr COFINS":           round(float(depois["Vlr COFINS"]) - _fv(antes.get("Vlr COFINS","0")), 2),
            "Tem Diferença":             "SIM" if tem_diff else "NÃO",
        }
        diferencas.append(row_conf)

    if not modificado:
        return None, "nenhuma alteração aplicada", "info", []

    recalcular_totais(inf_nfe)
    xml_out = etree.tostring(tree, xml_declaration=True, encoding="UTF-8", pretty_print=False)
    return xml_out, f"alterado com sucesso ({len(itens_validos)} itens)", "ok", diferencas

# ── Recalcula ICMSTot ──────────────────────────────────────────────────────────
def recalcular_totais(inf_nfe):
    totais = {k: 0.0 for k in ["vBC","vICMS","vBCST","vST","vIPI","vIPIDevol","vPIS","vCOFINS","vProd"]}
    for filho in inf_nfe:
        if local(filho) != "det": continue
        imposto = find(filho, "imposto"); prod = find(filho, "prod")
        if prod:
            el = find(prod, "vProd")
            if el is not None and el.text:
                try: totais["vProd"] += float(el.text)
                except Exception: pass
        if imposto:
            icms_pai = find(imposto, "ICMS")
            if icms_pai:
                for icms_f in icms_pai:
                    if local(icms_f).startswith("ICMS"):
                        for k, t in [("vBC","vBC"),("vICMS","vICMS"),("vBCST","vBCST"),("vICMSST","vST")]:
                            el = find(icms_f, k)
                            if el is not None and el.text:
                                try: totais[t] += float(el.text)
                                except Exception: pass
                        break
            ipi_trib = find(imposto, "IPI", "IPITrib")
            if ipi_trib:
                el = find(ipi_trib, "vIPI")
                if el is not None and el.text:
                    try: totais["vIPI"] += float(el.text)
                    except Exception: pass
            for tag, key in [("PIS","vPIS"), ("COFINS","vCOFINS")]:
                pai = find(imposto, tag)
                if pai:
                    for f in pai:
                        el = find(f, key)
                        if el is not None and el.text:
                            try: totais[key] += float(el.text)
                            except Exception: pass
                        break
        imp_devol = find(filho, "impostoDevol")
        if imp_devol:
            el = find(imp_devol, "IPI", "vIPIDevol")
            if el is not None and el.text:
                try: totais["vIPIDevol"] += float(el.text)
                except Exception: pass
    icms_tot = find(inf_nfe, "total", "ICMSTot")
    if icms_tot is None: return
    for tn, tv in {"vBC": fmt(totais["vBC"]), "vICMS": fmt(totais["vICMS"]),
                   "vBCST": fmt(totais["vBCST"]), "vST": fmt(totais["vST"]),
                   "vIPI": fmt(totais["vIPI"]), "vIPIDevol": fmt(totais["vIPIDevol"]),
                   "vPIS": fmt(totais["vPIS"]), "vCOFINS": fmt(totais["vCOFINS"]),
                   "vProd": fmt(totais["vProd"]),
                   "vNF": fmt(totais["vProd"] + totais["vST"] + totais["vIPI"])}.items():
        el = find(icms_tot, tn)
        if el is not None: el.text = tv

# ── Excel de Conferência ───────────────────────────────────────────────────────
COLUNAS_NUMERICAS = [
    "Vlr Documento",
    "BC ICMS Antes","BC ICMS Depois","% ICMS","Vlr ICMS Antes","Vlr ICMS Depois","Diff Vlr ICMS",
    "BC ICMS ST Antes","BC ICMS ST Depois","Vlr ICMS ST Antes","Vlr ICMS ST Depois","Diff Vlr ICMS ST",
    "BC IPI Antes","BC IPI Depois","% IPI","Vlr IPI Antes","Vlr IPI Depois","Diff Vlr IPI",
    "pDevol Antes","pDevol Depois","vIPIDevol Antes","vIPIDevol Depois","Diff vIPIDevol",
    "BC PIS Antes","BC PIS Depois","% PIS","Vlr PIS Antes","Vlr PIS Depois","Diff Vlr PIS",
    "BC COFINS Antes","BC COFINS Depois","% COFINS","Vlr COFINS Antes","Vlr COFINS Depois","Diff Vlr COFINS",
]

def gerar_excel_conferencia(todas_diferencas: list) -> bytes:
    if not todas_diferencas: return b""
    df = pd.DataFrame(todas_diferencas)
    for col in COLUNAS_NUMERICAS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Conferência")
        ws = writer.sheets["Conferência"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        COR_LARANJA="FF8000"; COR_CINZA="444444"; COR_BRANCO="FFFFFF"
        COR_VERDE_BG="E2EFDA"; COR_VERDE_FT="375623"
        COR_VERM_BG="FCE4D6"; COR_VERM_FT="843C0C"
        COR_PAR="F9F9F9"; COR_TOTAL_BG="444444"; COR_TOTAL_FT="FF8000"
        thin=Side(style="thin",color="CCCCCC"); thick=Side(style="medium",color=COR_LARANJA)
        b_hdr=Border(left=thick,right=thick,top=thick,bottom=thick)
        b_nrm=Border(left=thin,right=thin,top=thin,bottom=thin)
        b_tot=Border(left=thick,right=thick,top=thick,bottom=thick)
        grupos = {
            "id":     {"cols":["Arquivo","Chave NF-e","Nro Documento","Data Emissão","Data Entrada",
                               "Razão Social","CNPJ/CPF","UF","nItem","Cód Item","Desc Item","NCM",
                               "CFOP XML (origem)","CFOP XLSX (DNI/entrada)","Vlr Documento"],"cor":"2E4057"},
            "icms":   {"cols":["CST ICMS Antes","CST ICMS Depois","BC ICMS Antes","BC ICMS Depois",
                               "% ICMS","Vlr ICMS Antes","Vlr ICMS Depois","Diff Vlr ICMS",
                               "BC ICMS ST Antes","BC ICMS ST Depois","Vlr ICMS ST Antes",
                               "Vlr ICMS ST Depois","Diff Vlr ICMS ST"],"cor":"1F4E79"},
            "ipi":    {"cols":["CST IPI Antes","CST IPI Depois","BC IPI Antes","BC IPI Depois",
                               "% IPI","Vlr IPI Antes","Vlr IPI Depois","Diff Vlr IPI"],"cor":"375623"},
            "devol":  {"cols":["pDevol Antes","pDevol Depois","vIPIDevol Antes","vIPIDevol Depois",
                               "Diff vIPIDevol"],"cor":"7B3F00"},
            "pis":    {"cols":["CST PIS Antes","CST PIS Depois","BC PIS Antes","BC PIS Depois",
                               "% PIS","Vlr PIS Antes","Vlr PIS Depois","Diff Vlr PIS"],"cor":"7B2C2C"},
            "cofins": {"cols":["CST COFINS Antes","CST COFINS Depois","BC COFINS Antes","BC COFINS Depois",
                               "% COFINS","Vlr COFINS Antes","Vlr COFINS Depois","Diff Vlr COFINS"],"cor":"843C0C"},
            "flag":   {"cols":["Tem Diferença"],"cor":COR_LARANJA},
        }
        col_names = list(df.columns)
        col_grupo = {}
        for grp, info in grupos.items():
            for c in info["cols"]:
                if c in col_names:
                    col_grupo[col_names.index(c)+1] = info["cor"]
        for cell in ws[1]:
            cor = col_grupo.get(cell.column, COR_CINZA)
            cell.fill=PatternFill("solid",fgColor=cor)
            cell.font=Font(color=COR_BRANCO,bold=True,size=9,name="Segoe UI")
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=b_hdr
        ws.row_dimensions[1].height=36
        diff_cols=[i+1 for i,c in enumerate(col_names) if c.startswith("Diff ")]
        flag_col=col_names.index("Tem Diferença")+1 if "Tem Diferença" in col_names else None
        fill_par=PatternFill("solid",fgColor=COR_PAR)
        num_col_indices={col_names.index(c)+1:c for c in COLUNAS_NUMERICAS if c in col_names}
        NUM_FMT='#,##0.00'
        for row_idx, row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row),start=2):
            is_par=(row_idx%2==0)
            for cell in row:
                cell.border=b_nrm
                cell.alignment=Alignment(horizontal="center",vertical="center")
                cell.font=Font(size=9,name="Segoe UI")
                if is_par: cell.fill=fill_par
                if cell.column in num_col_indices:
                    cell.number_format=NUM_FMT
                    cell.alignment=Alignment(horizontal="right",vertical="center")
            for col_idx in diff_cols:
                cell=row[col_idx-1]
                try:
                    v=float(cell.value or 0)
                    if v>0:
                        cell.fill=PatternFill("solid",fgColor=COR_VERDE_BG)
                        cell.font=Font(color=COR_VERDE_FT,bold=True,size=9,name="Segoe UI")
                    elif v<0:
                        cell.fill=PatternFill("solid",fgColor=COR_VERM_BG)
                        cell.font=Font(color=COR_VERM_FT,bold=True,size=9,name="Segoe UI")
                except Exception: pass
            if flag_col:
                cell=row[flag_col-1]
                if str(cell.value)=="SIM":
                    cell.fill=PatternFill("solid",fgColor="FFE0B2")
                    cell.font=Font(color=COR_LARANJA,bold=True,size=9,name="Segoe UI")
                else:
                    cell.fill=PatternFill("solid",fgColor=COR_VERDE_BG)
                    cell.font=Font(color=COR_VERDE_FT,bold=True,size=9,name="Segoe UI")

        total_row_idx=ws.max_row+1
        fill_total=PatternFill("solid",fgColor=COR_TOTAL_BG)
        font_total=Font(color=COR_TOTAL_FT,bold=True,size=9,name="Segoe UI")
        font_total_w=Font(color=COR_BRANCO,bold=True,size=9,name="Segoe UI")
        first_data_row=2; last_data_row=ws.max_row
        for col_idx in range(1,ws.max_column+1):
            cell=ws.cell(row=total_row_idx,column=col_idx)
            cell.fill=fill_total; cell.border=b_tot
            cell.alignment=Alignment(horizontal="center",vertical="center")
            col_name=col_names[col_idx-1] if col_idx<=len(col_names) else ""
            if col_idx==1:
                cell.value="TOTAL"; cell.font=font_total
            elif col_idx in num_col_indices:
                col_letter=get_column_letter(col_idx)
                cell.value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
                cell.number_format=NUM_FMT; cell.font=font_total
                cell.alignment=Alignment(horizontal="right",vertical="center")
            elif col_name=="nItem":
                col_letter=get_column_letter(col_idx)
                cell.value=f"=COUNTA({col_letter}{first_data_row}:{col_letter}{last_data_row})"
                cell.font=font_total_w
            elif col_name=="Tem Diferença":
                col_letter=get_column_letter(col_idx)
                cell.value=f'=COUNTIF({col_letter}{first_data_row}:{col_letter}{last_data_row},"SIM")'
                cell.font=font_total
            else:
                cell.value=""; cell.font=font_total_w
        ws.row_dimensions[total_row_idx].height=20

        for col_idx in range(1,ws.max_column+1):
            col_letter=get_column_letter(col_idx)
            max_len=len(str(ws.cell(1,col_idx).value or ""))
            for r in range(2,min(last_data_row+2,300)):
                v=str(ws.cell(r,col_idx).value or "")
                max_len=max(max_len,len(v))
            ws.column_dimensions[col_letter].width=min(max_len+3,40)
        for cn in ["Desc Item","Razão Social","Arquivo","Chave NF-e","CFOP XML (origem)","CFOP XLSX (DNI/entrada)"]:
            if cn in col_names:
                ws.column_dimensions[get_column_letter(col_names.index(cn)+1)].width=30
        ws.freeze_panes="A2"
        ws.auto_filter.ref=f"A1:{get_column_letter(ws.max_column)}{last_data_row}"
    buf.seek(0)
    return buf.read()

# ── ZIP final ──────────────────────────────────────────────────────────────────
def gerar_zip_completo(xmls_modificados: dict, xmls_originais: dict) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        nomes_modificados = set(xmls_modificados.keys())
        for nome, conteudo in xmls_modificados.items():
            zf.writestr(f"modificados/{nome}", conteudo)
        for nome, conteudo in xmls_originais.items():
            if nome not in nomes_modificados:
                zf.writestr(f"nao_alterados/{nome}", conteudo)
    buf.seek(0)
    return buf.read()

# ── Resultados ─────────────────────────────────────────────────────────────────
def render_resultados():
    m = st.session_state.resultado_metricas
    if m is None: return

    st.divider()
    st.markdown("### ✅ Resultado do Processamento")
    col_np, _ = st.columns([2, 6])
    with col_np:
        if st.button("🔁 Iniciar Novo Processo", key="btn_novo_processo",
                     use_container_width=True, type="primary"):
            _limpar_resultados(); st.rerun()

    st.markdown("")
    col_x, col_y, col_z, col_w, col_v = st.columns(5)
    col_x.metric("✅ Alterados",        m["ok"])
    col_y.metric("ℹ️ Sem alteração",    m["info"])
    col_z.metric("❌ Erros",            m["erro"])
    col_w.metric("📋 Itens c/ diff",    m["diff"])
    col_v.metric("📦 Total de XMLs",    m["total_xmls"])

    st.divider()

    xmls_mod  = st.session_state.resultado_xmls_modificados
    xmls_orig = st.session_state.resultado_xmls_originais
    total_mod = len(xmls_mod)
    total_nao = len(xmls_orig) - total_mod

    col_dl1, col_dl2 = st.columns(2)

    with col_dl1:
        excel_bytes = st.session_state.resultado_excel_bytes
        diferencas  = st.session_state.resultado_diferencas
        if excel_bytes and len(excel_bytes) > 0:
            diff_c = sum(1 for r in diferencas if r.get("Tem Diferença") == "SIM")
            st.download_button(
                label=f"📥 Excel de Conferência ({len(diferencas)} itens / {diff_c} com diff)",
                data=excel_bytes,
                file_name="conferencia_nfe_dni.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_excel_resultado",
            )
        else:
            st.info("ℹ️ Nenhum item processado para conferência.")

    with col_dl2:
        if xmls_orig:
            zip_completo = gerar_zip_completo(xmls_mod, xmls_orig)
            st.download_button(
                label=(
                    f"⬇ Baixar ZIP completo "
                    f"({total_mod} modificado(s) + {total_nao} não alterado(s))"
                ),
                data=zip_completo,
                file_name="xmls_completo_dni.zip",
                mime="application/zip",
                use_container_width=True, key="dl_zip_completo",
            )
            if total_mod > 0:
                buf_mod = io.BytesIO()
                with zipfile.ZipFile(buf_mod, "w", zipfile.ZIP_DEFLATED) as zf:
                    for nome, conteudo in xmls_mod.items():
                        zf.writestr(nome, conteudo)
                buf_mod.seek(0)
                st.download_button(
                    label=f"⬇ Baixar apenas modificados ({total_mod} XML(s))",
                    data=buf_mod.read(),
                    file_name="xmls_modificados_dni.zip",
                    mime="application/zip",
                    use_container_width=True, key="dl_zip_modificados",
                )
        else:
            st.warning("⚠️ Nenhum XML disponível para download.")

    with st.expander("📋 Log detalhado", expanded=False):
        for nome_arq, msg, status in st.session_state.resultado_log:
            icon = "✅" if status == "ok" else ("ℹ️" if status == "info" else "❌")
            st.markdown(f"{icon} `{nome_arq}` — {msg}")

# ── Interface principal ────────────────────────────────────────────────────────
def main():
    _init_state()
    render_sidebar()

    cfops_ativas_xlsx: set[str] = {str(c) for c in st.session_state.get("cfop_ativos", ATIVOS_PADRAO)}

    st.markdown("""
    <div class="main-header">
        <h2>🧾 Enriquecedor de NF-e — DNI</h2>
        <p>Thomson Reuters · Domínio Sistemas · v8.0 · XLSX e CSV suportados · Encoding automático · IPI: IPINT→IPITrib corrigido</p>
    </div>
    """, unsafe_allow_html=True)

    st.info(
        "ℹ️ **Lógica De/Para de CFOP:** O CFOP do XLSX/CSV é usado **apenas para filtrar** itens de devolução. "
        "O **CFOP original do XML nunca é alterado**. "
        "O ZIP final contém **todos os XMLs**: modificados na pasta `modificados/` e "
        "os não alterados na pasta `nao_alterados/`."
    )

    if st.session_state.processamento_concluido:
        st.info("📌 Resultado disponível. Clique em **🔁 Iniciar Novo Processo** para processar novos arquivos.")
        render_resultados()
        st.markdown('<div class="footer">Thomson Reuters · Domínio Sistemas · Enriquecedor NF-e v8.0 · DNI</div>',
                    unsafe_allow_html=True)
        return

    col_up1, col_up2 = st.columns(2)
    with col_up1:
        st.markdown("#### 📂 Planilha XLSX ou CSV — Domínio Sistemas")
        arquivo_xlsx = st.file_uploader(
            "planilha", type=["xlsx", "csv"], key="xlsx", label_visibility="collapsed"
        )
    with col_up2:
        st.markdown("#### 📄 Arquivos XML de NF-e")
        arquivos_xml = st.file_uploader("xmls", type=["xml","zip"], accept_multiple_files=True,
                                        key="xmls", label_visibility="collapsed")

    lista_ativos_str = sorted(cfops_ativas_xlsx)
    st.info(
        f"🔢 **CFOPs de devolução ativos (Planilha):** "
        f"{', '.join(lista_ativos_str) if lista_ativos_str else '⚠️ Nenhum CFOP ativo — configure na barra lateral.'}"
    )
    st.divider()

    if st.button("▶ Processar XMLs", type="primary", key="btn_processar", use_container_width=True):
        if not arquivo_xlsx:      st.error("❌ Selecione a planilha (XLSX ou CSV)."); st.stop()
        if not arquivos_xml:      st.error("❌ Selecione ao menos um XML ou ZIP."); st.stop()
        if not cfops_ativas_xlsx: st.error("❌ Nenhum CFOP ativo. Configure na barra lateral."); st.stop()

        with st.spinner("🔄 Lendo planilha..."):
            dados_indexados = ler_planilha(arquivo_xlsx.read(), arquivo_xlsx.name)
        if not dados_indexados: st.error("❌ Nenhum item válido na planilha."); st.stop()

        col_a, col_b, col_c = st.columns(3)
        col_a.metric("📊 Itens indexados", len(dados_indexados))
        col_b.metric("🔢 CFOPs ativos",    len(cfops_ativas_xlsx))
        col_c.metric("📁 Arquivos",        len(arquivos_xml))

        with st.expander("🔍 Debug — primeiros 5 itens indexados"):
            for i, ((ch, seq), d) in enumerate(list(dados_indexados.items())[:5]):
                st.code(
                    f"Chave: {ch} | Seq: {seq}\n"
                    f"CFOP Planilha (filtro): {d['cfop_xlsx']}\n"
                    f"CST ICMS: {d['cst_icms']} | vICMS: {d['vlr_icms']} | "
                    f"BC ST: {d['base_icms_st']} | vST: {d['vlr_icms_st']}\n"
                    f"CST IPI: {d['cst_ipi']} | BC IPI: {d['base_ipi']} | "
                    f"% IPI: {d['perc_ipi']} | vIPI: {d['vlr_ipi']} | tem_ipi: {d['tem_ipi']}"
                )

        xmls_para_processar: dict[str, bytes] = {}
        for arq in arquivos_xml:
            if arq.name.lower().endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(arq.read())) as zf:
                    for nome in zf.namelist():
                        if nome.lower().endswith(".xml"):
                            xmls_para_processar[nome] = zf.read(nome)
            else:
                xmls_para_processar[arq.name] = arq.read()

        resultados        = []
        xmls_modificados  = {}
        todas_diferencas  = []
        progress = st.progress(0)
        total    = len(xmls_para_processar)

        for idx, (nome_arq, conteudo) in enumerate(xmls_para_processar.items()):
            xml_out, msg, status, diffs = processar_xml(
                conteudo, nome_arq, dados_indexados, cfops_ativas_xlsx)
            resultados.append((nome_arq, msg, status))
            if xml_out:
                xmls_modificados[nome_arq] = xml_out
            todas_diferencas.extend(diffs)
            progress.progress((idx + 1) / total)
        progress.empty()

        excel_bytes = gerar_excel_conferencia(todas_diferencas) if todas_diferencas else b""

        st.session_state.resultado_excel_bytes       = excel_bytes
        st.session_state.resultado_xmls_modificados  = xmls_modificados
        st.session_state.resultado_xmls_originais    = xmls_para_processar
        st.session_state.resultado_log               = resultados
        st.session_state.resultado_diferencas        = todas_diferencas
        st.session_state.resultado_metricas          = {
            "ok":         sum(1 for _, _, s in resultados if s == "ok"),
            "info":       sum(1 for _, _, s in resultados if s == "info"),
            "erro":       sum(1 for _, _, s in resultados if s == "erro"),
            "diff":       sum(1 for r in todas_diferencas if r.get("Tem Diferença") == "SIM"),
            "total_xmls": total,
        }
        st.session_state.processamento_concluido = True
        st.rerun()

    st.markdown('<div class="footer">Thomson Reuters · Domínio Sistemas · Enriquecedor NF-e v8.0 · DNI</div>',
                unsafe_allow_html=True)


if __name__ == "__main__":
    main()
